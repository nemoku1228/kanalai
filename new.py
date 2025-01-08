import pandas as pd
import openpyxl
from openpyxl import load_workbook
import subprocess
import time, traceback
import webbrowser
from bs4 import BeautifulSoup, NavigableString
import requests, random, json
import os, unidecode
from tkinter import *
from tkinter import ttk
from tkinter.messagebox import showinfo
import re
import math
import tkinter as tk
from tkinter import filedialog

# API URL
url = "https://api.eduide.cc/v1/chat/completions"

# Maksimalus simbolių skaičius vienoje žinutėje
MAX_LENGTH = 3500  # Kad API veiktų stabiliai

# Kalbos, į kurias reikia versti
languages = {
    "Lietuvių": "lt",
    "Latvių": "lv",
    "Estų": "et",
    "Rusų": "ru",
    "Suomių": "fi",
}

def clean_html_output(html_content):
    """Pašalina nereikalingas žymas ir suformatuoja HTML tekstą."""
    cleaned_content = re.sub(r"```html|```", "", html_content)
    cleaned_content = cleaned_content.strip()
    return cleaned_content
    
def clean_html(html_content):
    """Pašalina HTML žymes ir grąžina švarų tekstą."""
    soup = BeautifulSoup(html_content, "html.parser")
    return soup.get_text(separator="\n").strip()

def split_message(message, max_length):
    """Padalina žinutę į dalis, neviršijančias nustatyto ilgio."""
    parts = []
    while len(message) > max_length:
        split_index = message.rfind(" ", 0, max_length)
        if split_index == -1:
            split_index = max_length
        parts.append(message[:split_index].strip())
        message = message[split_index:].strip()
    parts.append(message)
    return parts

def send_to_api(session, part, language):
    """Siunčia dalį į API, apdoroja atsakymą ir grąžina švarų tekstą."""
    data = {
        "model": "gpt-4",
        "messages": [
            {
                "role": "system",
                "content": f"""Please translate the given text into {language}. Additionally:
1. Make the text concise and well-structured by removing unnecessary words, symbols, or excessive spacing.
2. Format the text in clean HTML, using only black font color and maintaining consistent, small line spacing between paragraphs.
3. Do not include images, large gaps, or complex styles in the output.
4. Ensure the text looks aesthetically pleasing and is easy to read.
5. If the text contains lists, format them as simple HTML lists, and highlight key terms or phrases where necessary.
6. Do not translate the text word-for-word. Ensure the translation is natural, logical, and flows well in {language}. Polish words or phrases must not appear in the translation. If necessary, restructure or rephrase the text to make it more coherent and suitable for the target language ({language}).
7. Replace title and make current title short, unique and understandable by their description"""
            },
            {"role": "user", "content": part}
        ],
        "max_tokens": 1500,
        "temperature": 0.2,
    }
    response = session.post(url, headers={
        "accept": "*/*",
        "content-type": "application/json"
    }, json=data)

    if response.status_code == 200:
        raw_content = response.json()["choices"][0]["message"]["content"]
        return clean_html_output(raw_content)  # Pašaliname nereikalingas žymas
    else:
        print(f"Klaida: {response.status_code}")
        print(response.text)
        return None

def process_translations_for_files(file_paths):
    """Apdoroja kelis HTML failus ir išsaugo vertimus pagal failo pavadinimą."""
    for file_path in file_paths:
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                html_content = file.read()

            # Gauti aplanko, kuriame yra failas, kelią
            output_folder = os.path.dirname(file_path)
            original_file_name = os.path.splitext(os.path.basename(file_path))[0]

            print(f"\nApdorojamas failas: {original_file_name}.")
            print("Tekstas bus išverstas į šias kalbas: Lietuvių, Latvių, Estų, Rusų, Suomių.")
            
            # Procesas: versti tekstą ir išsaugoti failus
            process_translations_for_file(html_content, output_folder, original_file_name)

        except FileNotFoundError:
            print(f"Klaida: Failas {file_path} nerastas.")

def process_translations_for_file(user_input, output_folder, original_file_name):
    """Apdoroja vieną failą ir išsaugo vertimus pagal originalaus failo pavadinimą."""
    clean_text = clean_html(user_input)
    message_parts = split_message(clean_text, MAX_LENGTH)

    with requests.Session() as session:
        for language, lang_code in languages.items():
            print(f"\nVerčiama į {language} kalbą...")
            translated_parts = []

            for i, part in enumerate(message_parts):
                print(f"Siunčiama dalis {i + 1}/{len(message_parts)}: {part[:50]}...")
                response = send_to_api(session, part, language)
                if response:
                    translated_parts.append(response)
                else:
                    print(f"Klaida apdorojant dalį {i + 1} {language} kalbai.")
            
            # Sujungti visas dalis į vieną tekstą
            full_translation = "\n".join(translated_parts)
            
            # Sukurti švarų HTML failą kiekvienai kalbai
            output_file = os.path.join(output_folder, f"{original_file_name}_{language}.html")
            with open(output_file, 'w', encoding='utf-8') as file:
                file.write("<!DOCTYPE html>")
                file.write("<html lang='en'><head><title>Vertimas</title>")
                file.write("""
                <style>
                    body {
                        font-family: Arial, sans-serif;
                        font-size: 14px;
                        line-height: 1.5;
                        color: black;
                        margin: 20px;
                        padding: 0;
                    }
                    h1 {
                        font-size: 18px;
                        margin-bottom: 10px;
                    }
                    p {
                        margin: 5px 0;
                    }
                </style>
                </head><body>
                """)
                
                # Įtraukti tvarkingai suformatuotą tekstą
                formatted_text = full_translation.replace("\n", "</p><p>")
                file.write(f"<p>{formatted_text}</p>")
                
                file.write("</body></html>")
            
            print(f"Vertimas į {language} kalbą išsaugotas: {output_file}")

def select_html_files():
    """Leidžia pasirinkti kelis HTML failus per GUI."""
    root = tk.Tk()
    root.withdraw()  # Paslepia pagrindinį langą
    file_paths = filedialog.askopenfilenames(
        title="Pasirinkite HTML failus",
        filetypes=[("HTML failai", "*.html *.htm"), ("Visi failai", "*.*")]
    )
    return file_paths

def main():
    """Pagrindinė programa."""
    print("Pasirinkite HTML failus per atsidariusį langą.")
    file_paths = select_html_files()

    if file_paths:
        print("\nHTML failai sėkmingai pasirinkti.")
        process_translations_for_files(file_paths)
        print("\nVisi vertimai buvo išsaugoti atitinkamuose .html failuose.")
    else:
        print("HTML failai nebuvo pasirinkti.")

if __name__ == "__main__":
    main()
    
def set_token():
    dire = os.path.dirname(os.path.abspath(__file__))
    token_path = os.path.join(dire, 'token.txt')
    
    if os.path.exists(token_path):
        with open(token_path, 'r') as arquivo:
            token = arquivo.readline().strip()
            # Įrašykite tokeną į Entry laukelį
            entry_4.delete(0, END)
            entry_4.insert(0, token)
    else:
        print("token.txt failas nerastas!")

def create_links_dic(filepath):
    with open(filepath, "r", encoding="utf-8") as arquivo:
        conteudo_html = arquivo.read()
    soup = BeautifulSoup(conteudo_html, "html.parser")

    # Keywords to locate the ID
    id_keywords = ["SKU", "EAN", "GTIN", "EAN(GTIN)"]
    product_id = None
    for keyword in id_keywords:
        tag_with_id = soup.find(string=lambda text: text and keyword in text)
        if tag_with_id:
            next_tag = tag_with_id.find_next()
            if next_tag and next_tag.name in ["p", "span", "div", "td"]:
                product_id = next_tag.text.strip()
                break
            parent = tag_with_id.find_parent("tr")
            if parent:
                next_cell = parent.find_all("td")
                if len(next_cell) > 1:
                    product_id = next_cell[1].text.strip()
                    break

    image_links = []
    all_images_section = soup.find(string=re.compile(r"Images", re.IGNORECASE))
    if all_images_section:
        parent_section = all_images_section.find_parent("section")
        if parent_section:
            links = parent_section.find_all("a", href=True)
            for link in links:
                href = link["href"]
                if href not in image_links:
                    image_links.append(href)

    image_links = list(set(image_links))

    price_keywords = ["price", "cena"]
    price = None
    for keyword in price_keywords:
        tag_with_price = soup.find(string=lambda text: text and keyword in text.lower())
        if tag_with_price:
            potential_price_tag = tag_with_price.find_next()
            if potential_price_tag:
                match = re.search(r'\d{1,3}([.,]\d{2})?', potential_price_tag.text)
                if match:
                    price = match.group(0).replace(",", ".")
                    break

    return [product_id, image_links, price]

def get_st(html):
    print("Translating  : ")
    def remove_emojis(string):
        padrao_emojis = re.compile("[" \
                                    "\U0001F600-\U0001F64F"  # emoticons
                                    "\U0001F300-\U0001F5FF"  # symbols & pictograms
                                    "\U0001F680-\U0001F6FF"  # transport & map symbols
                                    "\U0001F1E0-\U0001F1FF"  # flags (iOS)
                                    "\U00002702-\U000027B0"
                                    "\U000024C2-\U0001F251"
                                    "]+", flags=re.UNICODE)
        string_sem_emojis = re.sub(padrao_emojis, '', string)
        return string_sem_emojis

    tags = html.split('>')
    fin = ''
    for tag in tags:
        if '<div' not in tag and '<img' not in tag and '<a' not in tag and '</div' not in tag and '</a' not in tag:
            if len(tag) > 1:
                tg = remove_emojis(tag)
                fin += (tg + '>')
    texto_formatado = fin.replace('<', ' <').replace('>', '> ')
    sup = BeautifulSoup(texto_formatado, 'html.parser')
    for teg in sup.find_all():
        if teg.name not in ['ul', 'li']:
            teg.name = 'p'
    texto_formatado = str(sup.prettify())
    return texto_formatado

def gerand():
    numero_digitos = random.randint(11, 13)
    numero_aleatorio = ''.join([str(random.randint(0, 9)) for _ in range(numero_digitos)])
    return numero_aleatorio

def extract_title(filepath):
    with open(filepath, "r", encoding="utf-8") as file:
        content = file.read()

    soup = BeautifulSoup(content, "html.parser")
    title_section = soup.find("h1", string=re.compile(r"Title:", re.IGNORECASE))
    if title_section:
        title_tag = title_section.find_next("p")
        if title_tag:
            title = title_tag.get_text(strip=True)
            return title
        title = title_section.get_text(strip=True).replace("Title:", "").strip()
        if title:
            return title

    generic_title = soup.find("h1")
    if generic_title:
        title = generic_title.get_text(strip=True)
        if title:
            return title

    return "Title not found"

def extract_dimensions_and_weight(filepath):
    with open(filepath, "r", encoding="utf-8") as file:
        content = file.read()
    soup = BeautifulSoup(content, "html.parser")

    result = {"length": 0.5, "width": 0.5, "height": 0.5, "weight": 5}

    keywords = {
        "length": ["długość", "dł", "length"],
        "width": ["szerokość", "szer", "width"],
        "height": ["wysokość", "wys", "height"],
        "weight": ["waga", "weight"],
        "combined": ["dimensions", "size", "rozmiar", "wymiary", "dł x szer x wys"]
    }

    def extract_near_keyword(soup, keyword_list, unit_conversion=None):
        for keyword in keyword_list:
            for tag in soup.find_all(string=re.compile(keyword, re.IGNORECASE)):
                tag_text = tag.strip()
                numbers = re.findall(r"\d+[.,]?\d*", tag_text)
                if numbers:
                    value = float(numbers[0].replace(",", "."))
                    if unit_conversion and "mm" in tag_text.lower():
                        value /= 1000
                    elif unit_conversion and "cm" in tag_text.lower():
                        value /= 100
                    return value

                for sibling in tag.find_all_next(string=True, limit=5):
                    sibling_text = sibling.strip()
                    numbers = re.findall(r"\d+[.,]?\d*", sibling_text)
                    if numbers:
                        value = float(numbers[0].replace(",", "."))
                        if unit_conversion and "mm" in sibling_text.lower():
                            value /= 1000
                        elif unit_conversion and "cm" in sibling_text.lower():
                            value /= 100
                        return value
        return None

    def extract_weight(soup):
        for tag in soup.find_all(string=re.compile(r"\b(?:waga|weight)\b", re.IGNORECASE)):
            for sibling in tag.find_all_next(string=True, limit=1):
                sibling_text = sibling.strip()
                weight_numbers = re.findall(r"\d+[.,]?\d*", sibling_text)
                if weight_numbers:
                    weight = float(weight_numbers[0].replace(",", "."))
                    if "g" in sibling_text.lower() and "kg" not in sibling_text.lower():
                        weight /= 1000
                    return weight
        return None

    result["weight"] = extract_weight(soup) or result["weight"]

    for combined_keyword in keywords["combined"]:
        for tag in soup.find_all(string=re.compile(combined_keyword, re.IGNORECASE)):
            combined_text = tag + "".join(tag.find_all_next(string=True, limit=5))
            dimensions_match = re.search(r"(\d+)[^\d]+(\d+)[^\d]+(\d+)\s*(mm|cm|m)?", combined_text)
            if dimensions_match:
                unit = dimensions_match.group(4) or "mm"
                length = float(dimensions_match.group(1))
                width = float(dimensions_match.group(2))
                height = float(dimensions_match.group(3))

                conversion_factor = 1000 if unit == "mm" else 100 if unit == "cm" else 1
                result["length"] = length / conversion_factor
                result["width"] = width / conversion_factor
                result["height"] = height / conversion_factor

    for dimension, keyword_list in {"length": keywords["length"], "width": keywords["width"], "height": keywords["height"]}.items():
        if result[dimension] == 0.5:
            result[dimension] = extract_near_keyword(soup, keyword_list, unit_conversion=True) or result[dimension]

    for key in result:
        if result[key] and key == "weight":
            result[key] == 5
        if result[key] == 0:
            result[key] = 0.5

    return result
    
    def scrap_description(n_p, html):
        set_token()
        try:
            filepath = html
            global thefile
            thefile = filepath
            features()
            print("This is the features list ################# \n \n \n \n", list_feat, "\n \n \n \n")
            id1, images_links, price = create_links_dic(filepath)
            print("This is the SKU ################ \n \n \n \n", id1, "\n \n \n \n")
            print("This is the images urls ################# \n \n \n \n", images_links, "\n \n \n \n")
            print("This the original price ################################## \n", price, "\n \n \n \n")
            contador_list.append(1)
            try:
                with open(filepath, "r", encoding="utf-8") as arquivo:
                    conteudo_html = arquivo.read()
                soup = BeautifulSoup(conteudo_html, "html.parser")
                descriptio_original = soup.find(attrs={'data-box-name': 'Description card'})
                print("########################### extracted description : \n \n \n", descriptio_original, "\n \n \n")
                print("############################ get str description", get_st(str(descriptio_original)), "\n \n \n")
                descriptio_div = make_it_unique(get_st(str(descriptio_original)), "d")

                print("This is the original description ##################### \n", get_st(str(descriptio_div)), "\n \n \n \n")
                description_lt = traduzir_texto(get_st(str(descriptio_div)), 'lt').replace('ą', 'a').replace('ę', 'e').replace('į', 'i').replace('Ą', 'A').replace('Ę', 'E').replace('Į', 'I').replace('Ų', 'U')
                print("this is the Lithuanian description #################### \n", description_lt, "\n \n \n \n")
                description_lv = traduzir_texto(get_st(str(descriptio_div)), 'lv').replace('ą', 'a').replace('ę', 'e').replace('į', 'i').replace('Ą', 'A').replace('Ę', 'E').replace('Į', 'I').replace('Ų', 'U')
                print("this is the Latvian description #################### \n", description_lv, "\n \n \n \n")
                description_et = traduzir_texto(get_st(str(descriptio_div)), 'et').replace('ą', 'a').replace('ę', 'e').replace('į', 'i').replace('Ą', 'A').replace('Ę', 'E').replace('Į', 'I').replace('Ų', 'U')
                print("this is the Estonian description #################### \n", description_et, "\n \n \n \n")
                description_ru = traduzir_texto(get_st(str(descriptio_div)), 'ru').replace('ą', 'a').replace('ę', 'e').replace('į', 'i').replace('Ą', 'A').replace('Ę', 'E').replace('Į', 'I').replace('Ų', 'U')
                print("this is the Russian description #################### \n", description_ru, "\n \n \n \n")
                description_fi = traduzir_texto(get_st(str(descriptio_div)), 'fi').replace('ą', 'a').replace('ę', 'e').replace('į', 'i').replace('Ą', 'A').replace('Ę', 'E').replace('Į', 'I').replace('Ų', 'U')
                print("this is the Finnish description #################### \n", description_fi, "\n \n \n \n")

                title_original = extract_title(filepath)
                print("This is the original title ############# \n \n \n \n", title_original.strip(), "\n \n \n \n")
                real_title = make_it_unique(title_original, "t")
                print("This is the new version of the original title ############# \n \n \n \n", real_title.strip(), "\n \n \n \n")
                title_fi = translate_title(real_title.strip(), 'fi').replace('ą', 'a').replace('ę', 'e').replace('į', 'i').replace('Ą', 'A').replace('Ę', 'E').replace('Į', 'I').replace('Ų', 'U')
                print("This is the Finnish title ############# \n \n \n \n", title_fi, "\n \n \n \n")
                title_lt = translate_title(real_title.strip(), 'lt').replace('ą', 'a').replace('ę', 'e').replace('į', 'i').replace('Ą', 'A').replace('Ę', 'E').replace('Į', 'I').replace('Ų', 'U')
                print("This is the Lithuanian title ############# \n \n \n \n", title_lt, "\n \n \n \n")
                title_lv = translate_title(real_title.strip(), 'lv').replace('ą', 'a').replace('ę', 'e').replace('į', 'i').replace('Ą', 'A').replace('Ę', 'E').replace('Į', 'I').replace('Ų', 'U')
                print("This is the Latvian title ############# \n \n \n \n", title_lv, "\n \n \n \n")
                title_et = translate_title(real_title.strip(), 'et').replace('ą', 'a').replace('ę', 'e').replace('į', 'i').replace('Ą', 'A').replace('Ę', 'E').replace('Į', 'I').replace('Ų', 'U')
                print("This is the Estonian title ############# \n \n \n \n", title_et, "\n \n \n \n")
                title_ru = translate_title(real_title.strip(), 'ru').replace('ą', 'a').replace('ę', 'e').replace('į', 'i').replace('Ą', 'A').replace('Ę', 'E').replace('Į', 'I').replace('Ų', 'U')
                print("This is the Russian title ############# \n \n \n \n", title_ru, "\n \n \n \n")
                sellers_id = '9995218'
                urrl = 'https://pmpapi.pigugroup.eu/v3/sellers/9995218/product/import/execution'
                resposta1 = requests.post(urrl, headers={"Authorization": f"Pigu-mp {entry_4.get()}"})
                execution_id = resposta1.text.replace('{', '').replace('}', '').replace('"id":', '')
                url1 = f"https://pmpapi.pigugroup.eu/v3/sellers/product/import/execution/{execution_id}"
                print(execution_id)

                ean = gerand()
                global fake
                fake = ean
                category1 = entry_9.get().split()[-1].replace('[', '').replace(']', '')
                print("\n \n \n", category1, " \n \n type :", type(category1))
                print("execution_id :::::", execution_id)
                if category1 == "":
                    category1 = "10451"
                category = int(category1)
                print()
                print("this the list features ::::", list_feat, "\n \n \n \n")
                dimensions = extract_dimensions_and_weight(filepath)
                images_links = [i + '.jpg' for i in images_links]
                print("these are the dimensions ################## \n", dimensions, "\n \n \n \n")
                djson = {
                    "category_id": category,
                    "title": title_lt,
                    "title_lv": title_lv,
                    "title_ee": title_et,
                    "title_fi": title_fi,
                    "title_ru": title_ru,
                    "long_description": get_st(description_lt),
                    "long_description_lv": get_st(description_lv),
                    "long_description_ee": get_st(description_et),
                    "long_description_fi": get_st(description_fi),
                    "long_description_ru": get_st(description_ru),
                    "youtube_videos": [""],
                    "product_features": list_feat,
                    "images": images_links,
                    "modifications": [
                        {
                            "title": "modification title",
                            "sku": id1,
                            "eans": [ean],
                            "package_weight": float(dimensions.get("weight")),
                            "package_length": float(dimensions.get("length")),
                            "package_height": float(dimensions.get("height")),
                            "package_width": float(dimensions.get("width"))
                        }
                    ]
                }
                dadosjson = json.dumps(djson, ensure_ascii=True)
                print("The First Data that extracted and sent to the api ####################################### \n", dadosjson, "\n \n \n \n")
                resposta = requests.post(url1, data=dadosjson, headers={"Authorization": f"Pigu-mp {entry_4.get()}", 'Content-Type': 'application/json; charset=utf-8'})
                if resposta.status_code != 200:
                    print("\n \n \n \n ###################### error you are not reaching the api ############################# \n \n \n \n")
                else:
                    print("Success adding the product!")
            except Exception as err:
                traceback.print_exc()
            if n_p == '1':
                print("Save changes and check the gray window that was opened")
                offer()
        except Exception as e:
            traceback.print_exc()
            showinfo(message="An error occurred during the process, please contact the developer.")

    def offer():
        amt1 = entry_8.get()
        if amt1 == "":
            amt1 = "1000"
        amt = int(amt1)
        delh1 = entry_7.get()
        if delh1 == "":
            delh1 = "144"
        delh = int(delh1)
        print("type of delh", type(delh), "\n \n \n \n")
        url_ofer = "https://pmpapi.pigugroup.eu/v3/offers"
        ean = create_links_dic(thefile)[0]
        print("type of ean", type(ean), "\n \n \n \n")
        print("type of ean", type(fake), "\n \n \n \n")
        s = create_links_dic(thefile)[2]
        print("This the original price ################################## \n", s, "\n \n \n \n")
        n = float(s.replace(',', '.'))
        price = f"{(n + (n * float(entry_91.get()) / 100)):.2f}"[:-2] + '99'
        priceaf = f"{(n + (n * float(entry_92.get()) / 100)):.2f}"[:-2] + '99'
        print("\n \n \n price test", price, "\n \n \n")
        print("type of price", type(price), "\n \n \n")
        print("price after test", priceaf, "\n \n \n")
        data = {
            "EAN": fake,
            "SKU": ean,
            "Price After Discount": priceaf,
            "Price": price,
            "Delivery Time": delh,
            "Amount": amt
        }
        df = pd.DataFrame([data])
        book = load_workbook(output_file)
        with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False, header=writer.sheets["Sheet1"].max_row == 0, startrow=writer.sheets["Sheet1"].max_row)
        for app_name in ['pigu.lt', '220.lv', 'kaup24.ee', 'hobbyhall.fi']:
            dadosofer = {
                "seller_id": 9995218,
                "app_name": app_name,
                "sku": ean,
                "ean": fake,
                "delivery_hours": delh,
                "amount": amt,
                "sell_price": price,
                "sell_price_after_discount": priceaf,
                "status": "active"
            }
            dadosoffer = json.dumps(dadosofer)
            print("The second data that sent to the api  ####################################### \n", dadosoffer, "\n \n \n \n")
            resposta_ofer = requests.post(url_ofer, data=dadosoffer, headers={"Authorization": f"Pigu-mp {entry_4.get()}", 'Content-Type': 'application/json; charset=utf-8'})
            print("Final offer response for ", app_name, ' : ')
            print(resposta_ofer.status_code)
            if resposta_ofer.status_code != 201:
                print("\n \n \n \n this is the response", resposta_ofer.text, "\n \n \n \n")
        print("Success, but always check in console for errors or warnings")

    def atualizar(e):
        texto = unidecode.unidecode(entry_9.get())
        opcoes_filtradas = [opcao for opcao in lista_cat if testo.lower() in unidecode.unidecode(opcao).lower()]
        entry_9['values'] = opcoes_filtradas

    d_b = {
        'Apsauginės, dezinfekcinės, medicininės prekės [11345]': [],
        'Vaistinės prekės su pažeista pakuote [17786]': [],
        'Specialios apsaugos priemonės [17721]': [],
        'Vaikui ir mamai [17470]': []
    }

    def features():
        category = entry_9.get()
        if category not in d_b:
            print(f"No features found for category: {category}")
            return
        with open(thefile, "r", encoding="utf-8") as file:
            html_content = file.read()
        soup = BeautifulSoup(html_content, "html.parser")
        text = soup.get_text(separator="\n", strip=True)
        text = text[:9800]
        parts = [text[i:i+5000] for i in range(0, len(text), 5000)]
        features_list = d_b[category]

        with requests.Session() as session:
            for feature in features_list:
                i = 0
                for part in parts:
                    i += 1
                    print("working with part: ", i)
                    feature_value = send_to_api(session, part, feature)
                    if feature_value != "Nenurodyta":
                        break
                list_feat.append({"name": feature, "value": feature_value})

    def scrap_folder():
        folder_path = filedialog.askdirectory(title="Select Folder Containing HTML Files")
        if not folder_path:
            print("No folder selected.")
            return
        html_files = [file for file in os.listdir(folder_path) if file.endswith('.html')]
        if not html_files:
            print("No HTML files found in the selected folder.")
            return
        for html_file in html_files:
            file_path = os.path.join(folder_path, html_file)
            scrap_description('1',file_path)
        
    
    def get_token():
        with open(dire+'/token.txt', 'r') as arquivo:
            primeira_linha = arquivo.readline()
        entry_4.insert(0,primeira_linha)
    lista_cat=list(d_b.keys())
    root= Tk()
    root.geometry(f"600x600")
    root.title("HIBM")
    frame_4=Frame(root, padx=10, pady=10, bg="#7B68EE")
    frame_4.pack(fill=BOTH, expand=True)
    label_4=Label(frame_4, text="Token:", bg="#B0E0E6")
    label_4.pack(side=LEFT, fill=BOTH, expand=True)    
    frame_4_1=Frame(frame_4, padx=10, pady=10)
    frame_4_1.pack(side=LEFT, fill=BOTH, expand=True)  
    entry_4=Entry(frame_4_1)
    entry_4.pack(fill=BOTH, expand=True)
    frame_7=Frame(root, padx=10, pady=10, bg="#7B68EE")
    frame_7.pack(fill=BOTH, expand=True)
    label_7=Label(frame_7, text="Products Delevery hours :", bg="#B0E0E6")
    label_7.pack(side=LEFT, fill=BOTH, expand=True)    
    frame_7_1=Frame(frame_7, padx=10, pady=10)
    frame_7_1.pack(side=LEFT, fill=BOTH, expand=True)  
    entry_7=Entry(frame_7_1)
    entry_7.pack(fill=BOTH, expand=True)
    frame_8=Frame(root, padx=10, pady=10, bg="#7B68EE")
    frame_8.pack(fill=BOTH, expand=True)
    label_8=Label(frame_8, text="Products Stock Amount :", bg="#B0E0E6")
    label_8.pack(side=LEFT, fill=BOTH, expand=True)    
    frame_8_1=Frame(frame_8, padx=10, pady=10)
    frame_8_1.pack(side=LEFT, fill=BOTH, expand=True)  
    entry_8=Entry(frame_8_1)
    entry_8.pack(fill=BOTH, expand=True)
    frame_9=Frame(root, padx=10, pady=10, bg="#7B68EE")
    frame_9.pack(fill=BOTH, expand=True)
    label_9=Label(frame_9, text="Category Id :", bg="#B0E0E6")
    label_9.pack(side=LEFT, fill=BOTH, expand=True)    
    frame_9_1=Frame(frame_9, padx=10, pady=10)
    frame_9_1.pack(side=LEFT, fill=BOTH, expand=True)  
    entry_9=ttk.Combobox(frame_9_1, values=lista_cat, state='normal')
    entry_9.pack(fill=BOTH, expand=True)
    entry_9.bind('<KeyRelease>', atualizar)
    entry_9.bind('<<ComboboxSelected>>')
    frame_91=Frame(root, padx=10, pady=10, bg="#7B68EE")
    frame_91.pack(fill=BOTH, expand=True)
    label_91=Label(frame_91, text="Price (%) :", bg="#B0E0E6")
    label_91.pack(side=LEFT, fill=BOTH, expand=True)    
    frame_91_1=Frame(frame_91, padx=10, pady=10)
    frame_91_1.pack(side=LEFT, fill=BOTH, expand=True)  
    entry_91=Entry(frame_91_1)
    entry_91.pack(fill=BOTH, expand=True)
    frame_92=Frame(root, padx=10, pady=10, bg="#7B68EE")
    frame_92.pack(fill=BOTH, expand=True)
    label_92=Label(frame_92, text="Price with discount (%) :", bg="#B0E0E6")
    label_92.pack(side=LEFT, fill=BOTH, expand=True)    
    frame_92_1=Frame(frame_92, padx=10, pady=10)
    frame_92_1.pack(side=LEFT, fill=BOTH, expand=True)  
    entry_92=Entry(frame_92_1)
    entry_92.pack(fill=BOTH, expand=True)
    frame_botao=Frame(root, bg="#7B68EE", padx=10, pady=10)
    frame_botao.pack(fill=BOTH, expand=True)
    botao= Button(frame_botao, text="Add Products to PIGU", command= lambda: scrap_folder())
    botao.pack(fill=Y, expand=True)
    get_token()
    mainloop()              
main()

